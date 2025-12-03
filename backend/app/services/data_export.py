"""
Data Export Service for Vocalysis
Exports feature extraction data to Excel for model retraining and patent process
Includes accuracy metrics and F1 scores logging
"""

import os
import json
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

# Export directory
EXPORT_DIR = os.getenv("EXPORT_DIR", "/tmp/vocalysis_exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


class DataExportService:
    """Service for exporting voice analysis data to Excel"""
    
    def __init__(self):
        self.export_dir = EXPORT_DIR
        os.makedirs(self.export_dir, exist_ok=True)
    
    def export_features_to_excel(
        self,
        features_list: List[Dict[str, Any]],
        metadata_list: List[Dict[str, Any]] = None,
        filename: str = None
    ) -> str:
        """
        Export feature extraction data to Excel file
        
        Args:
            features_list: List of feature dictionaries from voice analysis
            metadata_list: Optional list of metadata (user_id, timestamp, etc.)
            filename: Optional custom filename
        
        Returns:
            Path to the exported Excel file
        """
        if not filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"vocalysis_features_{timestamp}.xlsx"
        
        filepath = os.path.join(self.export_dir, filename)
        
        # Create workbook
        wb = Workbook()
        
        # === Sheet 1: Raw Features ===
        ws_features = wb.active
        ws_features.title = "Raw Features"
        
        # Flatten features and create DataFrame
        flat_features = []
        for i, features in enumerate(features_list):
            flat = self._flatten_dict(features)
            if metadata_list and i < len(metadata_list):
                flat.update(metadata_list[i])
            flat["sample_index"] = i + 1
            flat_features.append(flat)
        
        if flat_features:
            df_features = pd.DataFrame(flat_features)
            
            # Write to sheet
            for r_idx, row in enumerate(dataframe_to_rows(df_features, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_features.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
        
        # === Sheet 2: Feature Statistics ===
        ws_stats = wb.create_sheet("Feature Statistics")
        
        if flat_features:
            # Calculate statistics for numeric columns
            numeric_cols = df_features.select_dtypes(include=['float64', 'int64']).columns
            stats_data = []
            
            for col in numeric_cols:
                stats_data.append({
                    "Feature": col,
                    "Mean": df_features[col].mean(),
                    "Std": df_features[col].std(),
                    "Min": df_features[col].min(),
                    "Max": df_features[col].max(),
                    "Median": df_features[col].median(),
                    "Count": df_features[col].count()
                })
            
            df_stats = pd.DataFrame(stats_data)
            
            for r_idx, row in enumerate(dataframe_to_rows(df_stats, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_stats.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
        
        # === Sheet 3: Indian Demographics Optimization ===
        ws_indian = wb.create_sheet("Indian Demographics")
        
        indian_params = [
            ["Parameter", "North India", "South India", "East India", "West India", "Description"],
            ["Pitch Range (Hz) - Male", "85-180", "90-190", "85-175", "88-185", "Fundamental frequency range for male speakers"],
            ["Pitch Range (Hz) - Female", "165-280", "170-290", "165-275", "168-285", "Fundamental frequency range for female speakers"],
            ["Speech Rate (syllables/sec)", "3.5-5.5", "3.8-5.8", "3.4-5.4", "3.6-5.6", "Average speech rate by region"],
            ["Pause Duration (ms)", "200-600", "180-550", "220-650", "190-580", "Average pause duration between phrases"],
            ["Formant F1 Range (Hz)", "300-900", "320-920", "290-880", "310-900", "First formant frequency range"],
            ["Formant F2 Range (Hz)", "850-2500", "880-2550", "830-2450", "860-2500", "Second formant frequency range"],
            ["Jitter Threshold (%)", "1.5", "1.4", "1.6", "1.5", "Normal jitter threshold"],
            ["Shimmer Threshold (%)", "4.0", "3.8", "4.2", "4.0", "Normal shimmer threshold"],
            ["HNR Threshold (dB)", "18", "19", "17", "18", "Minimum healthy HNR"],
            ["Languages", "Hindi, Punjabi, Urdu", "Tamil, Telugu, Kannada, Malayalam", "Bengali, Odia, Assamese", "Marathi, Gujarati, Konkani", "Primary languages by region"],
        ]
        
        for r_idx, row in enumerate(indian_params, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_indian.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        # === Sheet 4: Clinical Biomarkers Reference ===
        ws_biomarkers = wb.create_sheet("Clinical Biomarkers")
        
        biomarkers = [
            ["Biomarker", "Normal Range", "Anxiety Indicator", "Depression Indicator", "Stress Indicator", "Research Reference"],
            ["Pitch CV", "0.15-0.25", "> 0.25", "< 0.15", "> 0.30", "Cummins et al., 2015"],
            ["Speech Rate", "2.5-4.5 syl/s", "> 4.5 syl/s", "< 2.0 syl/s", "> 5.0 syl/s", "Low et al., 2011"],
            ["Jitter (relative)", "< 0.02", "> 0.025", "< 0.01", "> 0.03", "Scherer, 2003"],
            ["Shimmer (relative)", "< 0.06", "> 0.07", "> 0.08", "> 0.07", "Ozdas et al., 2004"],
            ["HNR (dB)", "> 15", "< 14", "< 12", "< 13", "Murphy et al., 2000"],
            ["Pause Ratio", "0.15-0.30", "< 0.15", "> 0.35", "< 0.12", "Cannizzaro et al., 2004"],
            ["Energy CV", "0.20-0.40", "> 0.45", "< 0.15", "> 0.50", "Alpert et al., 2001"],
            ["F0 Mean (Male)", "100-150 Hz", "> 160 Hz", "< 90 Hz", "> 170 Hz", "Scherer, 1986"],
            ["F0 Mean (Female)", "180-230 Hz", "> 250 Hz", "< 160 Hz", "> 260 Hz", "Scherer, 1986"],
        ]
        
        for r_idx, row in enumerate(biomarkers, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_biomarkers.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"Features exported to: {filepath}")
        
        return filepath
    
    def _flatten_dict(self, d: Dict, parent_key: str = '', sep: str = '_') -> Dict:
        """Flatten nested dictionary"""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                for i, item in enumerate(v):
                    if isinstance(item, dict):
                        items.extend(self._flatten_dict(item, f"{new_key}_{i}", sep=sep).items())
                    else:
                        items.append((f"{new_key}_{i}", item))
            else:
                items.append((new_key, v))
        return dict(items)


class AccuracyMetricsLogger:
    """Logger for accuracy metrics and F1 scores"""
    
    def __init__(self):
        self.metrics_file = os.path.join(EXPORT_DIR, "accuracy_metrics.json")
        self.metrics_excel = os.path.join(EXPORT_DIR, "accuracy_metrics.xlsx")
        self._load_metrics()
    
    def _load_metrics(self):
        """Load existing metrics from file"""
        if os.path.exists(self.metrics_file):
            try:
                with open(self.metrics_file, 'r') as f:
                    self.metrics = json.load(f)
            except:
                self.metrics = {"predictions": [], "summary": {}}
        else:
            self.metrics = {"predictions": [], "summary": {}}
    
    def _save_metrics(self):
        """Save metrics to file"""
        with open(self.metrics_file, 'w') as f:
            json.dump(self.metrics, f, indent=2, default=str)
    
    def log_prediction(
        self,
        prediction_id: str,
        user_id: str,
        predicted_class: str,
        confidence: float,
        probabilities: List[float],
        features: Dict[str, Any],
        actual_class: Optional[str] = None,
        demographic_region: Optional[str] = None
    ):
        """
        Log a prediction for accuracy tracking
        
        Args:
            prediction_id: Unique prediction ID
            user_id: User ID
            predicted_class: Predicted mental state (normal, anxiety, depression, stress)
            confidence: Prediction confidence (0-1)
            probabilities: Class probabilities [normal, anxiety, depression, stress]
            features: Extracted voice features
            actual_class: Ground truth label (if available from clinical validation)
            demographic_region: Indian demographic region (north, south, east, west)
        """
        entry = {
            "timestamp": datetime.utcnow().isoformat(),
            "prediction_id": prediction_id,
            "user_id": user_id,
            "predicted_class": predicted_class,
            "confidence": confidence,
            "probabilities": {
                "normal": probabilities[0] if len(probabilities) > 0 else 0,
                "anxiety": probabilities[1] if len(probabilities) > 1 else 0,
                "depression": probabilities[2] if len(probabilities) > 2 else 0,
                "stress": probabilities[3] if len(probabilities) > 3 else 0
            },
            "actual_class": actual_class,
            "demographic_region": demographic_region,
            "is_validated": actual_class is not None,
            "is_correct": predicted_class == actual_class if actual_class else None,
            "key_features": {
                "pitch_mean": features.get("pitch_mean"),
                "pitch_cv": features.get("pitch_cv"),
                "speech_rate": features.get("speech_rate"),
                "jitter_rel": features.get("jitter_rel"),
                "shimmer_rel": features.get("shimmer_rel"),
                "hnr_mean": features.get("hnr_mean"),
                "rms_mean": features.get("rms_mean"),
                "silence_ratio": features.get("silence_ratio")
            }
        }
        
        self.metrics["predictions"].append(entry)
        self._save_metrics()
        self._update_summary()
        
        logger.info(f"Logged prediction {prediction_id}: {predicted_class} (confidence: {confidence:.2f})")
    
    def _update_summary(self):
        """Update summary statistics"""
        predictions = self.metrics["predictions"]
        validated = [p for p in predictions if p["is_validated"]]
        
        if not predictions:
            return
        
        # Overall statistics
        self.metrics["summary"] = {
            "total_predictions": len(predictions),
            "validated_predictions": len(validated),
            "average_confidence": sum(p["confidence"] for p in predictions) / len(predictions),
            "last_updated": datetime.utcnow().isoformat()
        }
        
        # Class distribution
        class_counts = {"normal": 0, "anxiety": 0, "depression": 0, "stress": 0}
        for p in predictions:
            if p["predicted_class"] in class_counts:
                class_counts[p["predicted_class"]] += 1
        self.metrics["summary"]["class_distribution"] = class_counts
        
        # Accuracy metrics (if validated data available)
        if validated:
            correct = sum(1 for p in validated if p["is_correct"])
            self.metrics["summary"]["accuracy"] = correct / len(validated)
            
            # Per-class metrics
            classes = ["normal", "anxiety", "depression", "stress"]
            class_metrics = {}
            
            for cls in classes:
                tp = sum(1 for p in validated if p["predicted_class"] == cls and p["actual_class"] == cls)
                fp = sum(1 for p in validated if p["predicted_class"] == cls and p["actual_class"] != cls)
                fn = sum(1 for p in validated if p["predicted_class"] != cls and p["actual_class"] == cls)
                
                precision = tp / (tp + fp) if (tp + fp) > 0 else 0
                recall = tp / (tp + fn) if (tp + fn) > 0 else 0
                f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
                
                class_metrics[cls] = {
                    "precision": precision,
                    "recall": recall,
                    "f1_score": f1,
                    "support": sum(1 for p in validated if p["actual_class"] == cls)
                }
            
            self.metrics["summary"]["class_metrics"] = class_metrics
            
            # Macro F1 score
            macro_f1 = sum(m["f1_score"] for m in class_metrics.values()) / len(classes)
            self.metrics["summary"]["macro_f1_score"] = macro_f1
            
            # Weighted F1 score
            total_support = sum(m["support"] for m in class_metrics.values())
            if total_support > 0:
                weighted_f1 = sum(m["f1_score"] * m["support"] for m in class_metrics.values()) / total_support
                self.metrics["summary"]["weighted_f1_score"] = weighted_f1
        
        # Regional metrics
        regional_counts = {}
        for p in predictions:
            region = p.get("demographic_region", "unknown")
            if region not in regional_counts:
                regional_counts[region] = {"count": 0, "correct": 0}
            regional_counts[region]["count"] += 1
            if p.get("is_correct"):
                regional_counts[region]["correct"] += 1
        
        for region, counts in regional_counts.items():
            if counts["count"] > 0:
                counts["accuracy"] = counts["correct"] / counts["count"] if counts["correct"] > 0 else None
        
        self.metrics["summary"]["regional_metrics"] = regional_counts
        
        self._save_metrics()
    
    def export_metrics_to_excel(self) -> str:
        """Export all metrics to Excel file"""
        wb = Workbook()
        
        # === Sheet 1: Predictions Log ===
        ws_predictions = wb.active
        ws_predictions.title = "Predictions Log"
        
        if self.metrics["predictions"]:
            # Flatten predictions for DataFrame
            flat_predictions = []
            for p in self.metrics["predictions"]:
                flat = {
                    "timestamp": p["timestamp"],
                    "prediction_id": p["prediction_id"],
                    "user_id": p["user_id"],
                    "predicted_class": p["predicted_class"],
                    "confidence": p["confidence"],
                    "prob_normal": p["probabilities"]["normal"],
                    "prob_anxiety": p["probabilities"]["anxiety"],
                    "prob_depression": p["probabilities"]["depression"],
                    "prob_stress": p["probabilities"]["stress"],
                    "actual_class": p["actual_class"],
                    "demographic_region": p["demographic_region"],
                    "is_validated": p["is_validated"],
                    "is_correct": p["is_correct"]
                }
                # Add key features
                for k, v in p.get("key_features", {}).items():
                    flat[f"feature_{k}"] = v
                flat_predictions.append(flat)
            
            df = pd.DataFrame(flat_predictions)
            
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_predictions.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
        
        # === Sheet 2: Summary Statistics ===
        ws_summary = wb.create_sheet("Summary Statistics")
        
        summary = self.metrics.get("summary", {})
        summary_rows = [
            ["Metric", "Value"],
            ["Total Predictions", summary.get("total_predictions", 0)],
            ["Validated Predictions", summary.get("validated_predictions", 0)],
            ["Average Confidence", f"{summary.get('average_confidence', 0):.4f}"],
            ["Overall Accuracy", f"{summary.get('accuracy', 'N/A')}"],
            ["Macro F1 Score", f"{summary.get('macro_f1_score', 'N/A')}"],
            ["Weighted F1 Score", f"{summary.get('weighted_f1_score', 'N/A')}"],
            ["Last Updated", summary.get("last_updated", "N/A")]
        ]
        
        for r_idx, row in enumerate(summary_rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        # === Sheet 3: Class Metrics ===
        ws_class = wb.create_sheet("Class Metrics")
        
        class_metrics = summary.get("class_metrics", {})
        class_rows = [["Class", "Precision", "Recall", "F1 Score", "Support"]]
        
        for cls, metrics in class_metrics.items():
            class_rows.append([
                cls.capitalize(),
                f"{metrics['precision']:.4f}",
                f"{metrics['recall']:.4f}",
                f"{metrics['f1_score']:.4f}",
                metrics['support']
            ])
        
        for r_idx, row in enumerate(class_rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_class.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        # === Sheet 4: Regional Metrics ===
        ws_regional = wb.create_sheet("Regional Metrics")
        
        regional_metrics = summary.get("regional_metrics", {})
        regional_rows = [["Region", "Count", "Correct", "Accuracy"]]
        
        for region, metrics in regional_metrics.items():
            regional_rows.append([
                region.capitalize() if region else "Unknown",
                metrics['count'],
                metrics.get('correct', 0),
                f"{metrics.get('accuracy', 'N/A')}"
            ])
        
        for r_idx, row in enumerate(regional_rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_regional.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        # Save workbook
        wb.save(self.metrics_excel)
        logger.info(f"Metrics exported to: {self.metrics_excel}")
        
        return self.metrics_excel
    
    def get_summary(self) -> Dict[str, Any]:
        """Get current summary statistics"""
        return self.metrics.get("summary", {})
    
    def validate_prediction(self, prediction_id: str, actual_class: str):
        """Add ground truth label to a prediction for accuracy calculation"""
        for p in self.metrics["predictions"]:
            if p["prediction_id"] == prediction_id:
                p["actual_class"] = actual_class
                p["is_validated"] = True
                p["is_correct"] = p["predicted_class"] == actual_class
                break
        
        self._save_metrics()
        self._update_summary()


# Singleton instances
data_export_service = DataExportService()
accuracy_logger = AccuracyMetricsLogger()
